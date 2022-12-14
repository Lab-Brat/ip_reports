import ipaddress
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict
from dataclasses import dataclass, fields

@dataclass
class Reader:
    ip_adrs: str
    prev_rep: str
    curr_rep: str
    clf: str
    legend: str

    def __post_init__(self):
        self.ip_adrs = self._read_txt(self.ip_adrs)
        self.prev_rep = self._xlsx_to_dict(self.prev_rep, 'Pentest Report')
        self.curr_rep = self.curr_rep
        self.clf = self._read_txt(self.clf)
        self.legend = self._read_txt(self.legend)

    def _read_txt(self, txt):
        '''
        read txt file line by line, output a list of lines
        '''
        with open(txt) as file:
            sn = [str(line.rstrip('\n')) for line in file]
        return sn

    def _is_ip(self, ip):
        '''
        determine if a string is a valid IP address
        '''
        try:
            ipaddress.IPv4Network(ip)
            return True
        except ValueError:
            return False

    def _determinte_color(self, sheet, cell):
        '''
        find cell color, only recognizes green, red, yellow
        '''
        color = sheet[cell].fill.start_color.index
        return color

    def _xlsx_to_dict(self, previous_report, sheet_name):
        '''
        read the previous report into a dictionary,
        format: {IP: [(port1, color1), (port1, color1), ...]...}
        '''
        prev_dict = defaultdict(list)
        sheet = load_workbook(previous_report, data_only=True)[sheet_name]

        for ip, port in zip(sheet['A'], sheet['B']):
            if self._is_ip(ip.value) == True:
                color = self._determinte_color(sheet, ip.coordinate)
                prev_dict[ip.value].append((port.value, color))

        return prev_dict

    def __repr__(self):
        cls = self.__class__
        cls_name = cls.__name__
        indent = ' ' * 4
        res = [f'{cls_name}(']
        for f in fields(cls):
            value = getattr(self, f.name)
            res.append(f'{indent}{f.name} ===> '
                       f'type: {type(value)}, entries: {len(value)},')
        res.append(')')
        return '\n'.join(res)


class Tools():
    def __init__(self, Reader):
        # Reader object
        self.reader = Reader

        # Colors
        self.colors = {'green': 'FF00FF00', 'red': 'FFFF0000',
                       'yellow': 'FFFFFF00', 'orange': 'FFFF6600',
                       'blue': 'FF00FFFF', 'gray': 'FFC0C0C0'}
    
    def _ip_to_digit(self, ip):
        '''
        convert IP address string to a unique integer
        '''
        return sum([int(oct) * (256 ** (3-i)) 
                    for i, oct in enumerate(str(ip).split('.'))])

    def subnet_range(self, subnet):
        '''
        get the 1st and last addresses from a subnet,
        return a tuple with the digit form of those 2 address
        '''
        net = ipaddress.IPv4Network(subnet)
        return (self._ip_to_digit(net[0]), self._ip_to_digit(net[-1]))

    def _is_ip(self, ip):
        '''
        determine if a string is a valid IP address
        '''
        try:
            ipaddress.IPv4Network(ip)
            return True
        except ValueError:
            return False

    def _is_present(self, ip, port):
        '''
        determine if a key is present in a dictionary
        '''
        try:
            for record in self.reader.prev_rep[ip]:
                if record[0] == port:
                    return record
            return False
        except KeyError:
            return False

    def _ip_within_range(self, ip, ranges):
        '''
        determine if an IP address belongs to a subnets
        '''
        for range in ranges:
            ipdig = self._ip_to_digit(ip)
            if range[0] <= ipdig <= range[1]:
                return True
        return False

    def _color_cells(self, row, color):
        '''
        loop through all cells in a row and color them
        '''
        for cell in row:
            cell.fill = PatternFill(start_color = color, 
                                    end_color = color,
                                    fill_type = "solid")

    def color_xlsx(self):
        '''
        read <self.crcsv> line by line and color accordingly
        '''
        workbook = load_workbook(self.reader.curr_rep, data_only=True)
        worksheet = workbook['Pentest Report']
        cf_ip_ranges = [self.subnet_range(ip) for ip in self.reader.clf]

        for row in worksheet:
            ip, port = str(row[0].value), str(row[1].value)
            record = self._is_present(ip, port)
            if self._is_ip(ip) == False:
                continue
            elif self._ip_within_range(ip, cf_ip_ranges):
                self._color_cells(row, self.colors['orange'])
            elif ip not in self.reader.ip_adrs:
                self._color_cells(row, self.colors['gray'])
            elif record:
                self._color_cells(row, record[1])
            else:
                self._color_cells(row, self.colors['blue'])
        workbook.save(self.reader.curr_rep)
